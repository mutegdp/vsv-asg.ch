from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup

from email_parser import email_parser


def export_xls(location, list_data):
    print("Saving the data into spreadsheets...")
    workbook = Workbook()
    sheet = workbook.active
    sheet.freeze_panes = "B2"
    sheet.cell(row=1, column=1).value = "Firmenname"
    sheet.cell(row=1, column=2).value = "Kontaktperson"
    sheet.cell(row=1, column=3).value = "Telefon"
    sheet.cell(row=1, column=4).value = "E-Mail"
    sheet.cell(row=1, column=5).value = "Private Equity?"
    sheet.column_dimensions["A"].width = 25
    sheet.column_dimensions["B"].width = 25
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 25

    for data in list_data:
        sheet.append(list(data.values()))

    workbook.save(location)
    print("Spreadsheet saved!")


def ada_nggak(list_data, cari, yes_no=False):
    for data in list_data:
        if yes_no:
            ada = [x for x in data if cari in x.lower()]
            if ada:
                return "Yes"
            else:
                continue
        if data[0].lower() == cari:
            return data[1]
        else:
            continue
    if yes_no:
        return "No"
    return ""


def get_data(_range=[]):
    target = "https://www.vsv-asg.ch/de/mitgliedersuche?typ=list"
    r = requests.get(target)
    soup = BeautifulSoup(r.content, "lxml")
    urls = soup.select("table.mitgliederliste tr td:nth-of-type(2) > a:nth-of-type(1)")

    all_data = []
    if _range:
        urls = urls[_range[0]:_range[1]]
    for idx, url in enumerate(urls, start=1):
        url = "https://www.vsv-asg.ch" + url.get("href")
        r = requests.get(url)
        print(f"getting...{idx}/{len(urls)} {url}")
        soup = BeautifulSoup(r.content, "lxml")
        table_data = soup.select("table.mitgliedersuche tr")

        page_data = []
        for _data in table_data:
            _data = _data.select("td")
            data = []
            for d in _data:
                if d.find("li"):
                    d = d.select("li")
                    d = [x.get_text().strip() for x in d]
                    d = ", ".join(d)
                else:
                    d = d.get_text().strip()
                if "*/eval(" in d:
                    d = email_parser.add(d)
                if d:
                    data.append(d)
            if data:
                page_data.append(data)

        company = {}
        company["Firmenname"] = ada_nggak(page_data, "firmenname")
        company["Kontaktperson"] = ada_nggak(page_data, "kontaktperson")
        company["Telefon"] = ada_nggak(page_data, "telefon")
        company["E-Mail"] = ada_nggak(page_data, "e-mail")
        company["Private Equity?"] = ada_nggak(page_data, "private equity", yes_no=True)
        all_data.append(company)

    return all_data
